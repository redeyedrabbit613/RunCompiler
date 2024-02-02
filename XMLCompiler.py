# -*- coding: utf-8 -*-
"""
Created on Thursday February 2 20:51:19 2024

@author: John Wang
"""
import glob as gl
import xml.etree.ElementTree as ET
import openpyxl as op
# Create a new workbook
workbook = op.Workbook()
# Select the active worksheet
worksheet = workbook.active


# Define the pattern to search for .XML files
pattern = ''

# Use glob to find the files matching the pattern
xml_files = gl.glob(pattern)

# Function to read attribute for a specific value
def read_attribute_for_value(xml_file, element_tag, attribute_name, attribute_name2, target_value):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    
    # Iterate over elements with the specified tag
    for elem in root.iter(element_tag):
        # Check if the element has the target value for the specified attribute
        if elem.get(attribute_name) == target_value:
            # Get the value of attribute_name2
            run_folder = root.attrib[attribute_name2]
            # Write the value to the next empty cell in the first column of the worksheet
            next_cell = worksheet.cell(row=worksheet.max_row + 1, column=1)
            next_cell.value = run_folder
            # Save the workbook to a file
            workbook.save('/path/to/directory/SOPRuns.xlsx')
            
# Specify the element tag, attribute name, attribute name 2, and target value
element_tag = 'ClusterRunMetadata'
attribute_name = 'Fragment_FilePath'
attribute_name2 = 'Run_Folder_Name'
target_value = 'STANDARD'




        
"""
Spyder Editor

This is a temporary script file.
"""

